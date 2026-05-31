"""
Geração de relatórios Excel (openpyxl) e PDF (reportlab).
Todas as funções retornam BytesIO prontos para streaming HTTP.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Dict, List, Optional

from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload

from app.models.area import Area
from app.models.ciclo import Ciclo
from app.models.lotacao import Lotacao
from app.models.procurador import Procurador
from app.models.vaga import Vaga

# ── cores por tipo de vaga (Excel ARGB) ──────────────────────────────────────

_VAGA_FILL: Dict[str, str] = {
    "PG":         "FFD1D5DB",  # cinza
    "NOMEACAO":   "FFFECACA",  # vermelho claro
    "ESCOLHA":    "FFBBF7D0",  # verde claro
    "DESIGNACAO": "FFFEF08A",  # amarelo claro
    "ACERVO":     "FFbfdbfe",  # azul claro
}

_TIPO_LABEL: Dict[str, str] = {
    "PG": "PG", "NOMEACAO": "Nomeação", "ESCOLHA": "Escolha",
    "DESIGNACAO": "Designação", "ACERVO": "Acervo",
}

_MOTIVO_LABEL: Dict[str, str] = {
    "POSSE_INICIAL": "Posse Inicial",
    "NOMEACAO": "Nomeação",
    "ESCOLHA_CHEFE": "Escolha de Chefe",
    "DESIGNACAO_PG": "Designação PG",
    "ACERVO": "Acervo",
    "PERMANENCIA": "Permanência",
    "SUBSTITUICAO_MANUAL": "Substituição Manual",
}


def _get_prev_areas(proc_ids: List[int], ciclo_id: str, db: Session) -> Dict[int, str]:
    """Retorna {proc_id: area_codigo} com a lotação mais recente ANTES do ciclo atual."""
    if not proc_ids:
        return {}
    prev_subq = (
        db.query(
            Lotacao.procurador_id,
            func.max(Lotacao.data_entrada).label("max_entrada"),
        )
        .filter(Lotacao.procurador_id.in_(proc_ids), Lotacao.ciclo_id != ciclo_id)
        .group_by(Lotacao.procurador_id)
        .subquery()
    )
    prev_lots = (
        db.query(Lotacao)
        .join(
            prev_subq,
            (Lotacao.procurador_id == prev_subq.c.procurador_id)
            & (Lotacao.data_entrada == prev_subq.c.max_entrada),
        )
        .filter(Lotacao.ciclo_id != ciclo_id)
        .all()
    )
    return {lot.procurador_id: lot.area_codigo for lot in prev_lots}


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def _wb_styles():
    """Importa estilos openpyxl (lazy para não quebrar testes sem dep)."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    return PatternFill, Font, Alignment, Border, Side


def gerar_mapa_xlsx(ciclo_id: str, db: Session) -> io.BytesIO:
    import openpyxl
    PatternFill, Font, Alignment, Border, Side = _wb_styles()

    ciclo = db.get(Ciclo, ciclo_id)
    if not ciclo:
        raise ValueError(f"Ciclo '{ciclo_id}' não encontrado")

    vagas = db.query(Vaga).filter(Vaga.ciclo_id == ciclo_id).order_by(
        Vaga.area_codigo, Vaga.tipo, Vaga.numero).all()
    procs = {p.id: p for p in db.query(Procurador).all()}
    areas = {a.codigo: a for a in db.query(Area).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mapa {ciclo_id}"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # cabeçalho
    headers = ["Área", "Nome da Área", "Tipo", "Nº", "Tipo Vaga", "Cargo", "Procurador", "Antiguidade", "Origem"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    row = 2
    for v in vagas:
        area = areas.get(v.area_codigo)
        proc = procs.get(v.ocupante_id) if v.ocupante_id else None
        fill = PatternFill("solid", fgColor=_VAGA_FILL.get(v.tipo, "FFFFFFFF"))

        values = [
            v.area_codigo,
            area.nome if area else "",
            area.tipo if area else "",
            v.numero,
            _TIPO_LABEL.get(v.tipo, v.tipo),
            v.cargo or "",
            proc.nome if proc else "",
            proc.antiguidade if proc else "",
            v.origem,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = thin
        row += 1

    # larguras automáticas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_lotacoes_xlsx(ciclo_id: str, db: Session) -> io.BytesIO:
    import openpyxl
    PatternFill, Font, Alignment, Border, Side = _wb_styles()

    ciclo = db.get(Ciclo, ciclo_id)
    if not ciclo:
        raise ValueError(f"Ciclo '{ciclo_id}' não encontrado")

    lotacoes = (
        db.query(Lotacao)
        .options(selectinload(Lotacao.procurador), selectinload(Lotacao.area))
        .filter(Lotacao.ciclo_id == ciclo_id)
        .order_by(Lotacao.area_codigo)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Lotações {ciclo_id}"

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Procurador", "Antiguidade", "Área", "Nome da Área", "Entrada", "Saída", "Motivo"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
        cell.border = thin

    for row, l in enumerate(lotacoes, 2):
        values = [
            l.procurador.nome if l.procurador else "",
            l.procurador.antiguidade if l.procurador else "",
            l.area_codigo,
            l.area.nome if l.area else "",
            l.data_entrada.isoformat() if l.data_entrada else "",
            l.data_saida.isoformat() if l.data_saida else "",
            l.motivo,
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val).border = thin

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_procuradores_xlsx(db: Session) -> io.BytesIO:
    import openpyxl
    PatternFill, Font, Alignment, Border, Side = _wb_styles()

    procs = db.query(Procurador).order_by(Procurador.antiguidade).all()
    areas = {a.codigo: a.nome for a in db.query(Area).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procuradores"

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Antiguidade", "Nome", "Status", "Lotação Atual", "Nome da Área", "Ativo"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
        cell.border = thin

    for row, p in enumerate(procs, 2):
        values = [
            p.antiguidade, p.nome, p.status,
            p.lotacao_atual_codigo or "",
            areas.get(p.lotacao_atual_codigo or "", ""),
            "Sim" if p.ativo else "Não",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val).border = thin

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_areas_xlsx(db: Session) -> io.BytesIO:
    import openpyxl
    PatternFill, Font, Alignment, Border, Side = _wb_styles()

    areas = db.query(Area).order_by(Area.tipo, Area.codigo).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Áreas"

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Código", "Nome", "Tipo", "PG", "Nomeação", "Escolha", "Designação", "Acervo", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
        cell.border = thin

    for row, a in enumerate(areas, 2):
        values = [
            a.codigo, a.nome, a.tipo,
            a.vagas_pg, a.vagas_nomeacao, a.vagas_escolha,
            a.vagas_designacao, a.vagas_acervo, a.total_vagas,
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val).border = thin

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# PDF (reportlab)
# ═══════════════════════════════════════════════════════════════════════════════

def _rl():
    """Importa reportlab lazy."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    return A4, colors, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, getSampleStyleSheet


_PDF_TIPO_COLOR = {
    "PG":         (0.82, 0.84, 0.86),
    "NOMEACAO":   (0.99, 0.79, 0.79),
    "ESCOLHA":    (0.73, 0.97, 0.81),
    "DESIGNACAO": (0.99, 0.94, 0.54),
    "ACERVO":     (0.75, 0.85, 0.99),
}


def gerar_mapa_pdf(ciclo_id: str, db: Session) -> io.BytesIO:
    A4, colors, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, getSS = _rl()

    ciclo = db.get(Ciclo, ciclo_id)
    if not ciclo:
        raise ValueError(f"Ciclo '{ciclo_id}' não encontrado")

    vagas = db.query(Vaga).filter(Vaga.ciclo_id == ciclo_id).order_by(
        Vaga.area_codigo, Vaga.tipo, Vaga.numero).all()
    procs = {p.id: p for p in db.query(Procurador).all()}
    areas = {a.codigo: a for a in db.query(Area).all()}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSS()
    story = []

    story.append(Paragraph(f"<b>Mapa de Relotação — Ciclo {ciclo_id}</b>", styles["Title"]))
    story.append(Paragraph(
        f"Gerado em {date.today().strftime('%d/%m/%Y')} · "
        f"Abertura: {ciclo.abertura.strftime('%d/%m/%Y') if ciclo.abertura else '—'}",
        styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    # Agrupa por área
    vagas_por_area: Dict[str, List[Vaga]] = {}
    for v in vagas:
        vagas_por_area.setdefault(v.area_codigo, []).append(v)

    for cod, vs in vagas_por_area.items():
        area = areas.get(cod)
        story.append(Paragraph(f"<b>{cod}</b> — {area.nome if area else ''}", styles["Heading2"]))

        data = [["Nº", "Tipo", "Cargo", "Procurador", "Antiguidade"]]
        row_types: List[str] = []
        for v in vs:
            proc = procs.get(v.ocupante_id) if v.ocupante_id else None
            data.append([
                str(v.numero),
                _TIPO_LABEL.get(v.tipo, v.tipo),
                v.cargo or "",
                proc.nome if proc else "—",
                str(proc.antiguidade) if proc else "",
            ])
            row_types.append(v.tipo)

        col_widths = [1*cm, 2.5*cm, 3*cm, 8*cm, 2*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ]
        for i, tipo in enumerate(row_types, 1):
            r, g, b = _PDF_TIPO_COLOR.get(tipo, (1, 1, 1))
            style_cmds.append(("BACKGROUND", (0, i), (1, i), colors.Color(r, g, b)))

        t.setStyle(TableStyle(style_cmds))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    buf.seek(0)
    return buf


def gerar_movimentacoes_xlsx(ciclo_id: str, db: Session) -> io.BytesIO:
    import openpyxl
    PatternFill, Font, Alignment, Border, Side = _wb_styles()

    ciclo = db.get(Ciclo, ciclo_id)
    if not ciclo:
        raise ValueError(f"Ciclo '{ciclo_id}' não encontrado")

    new_lots = (
        db.query(Lotacao)
        .options(selectinload(Lotacao.procurador))
        .filter(Lotacao.ciclo_id == ciclo_id)
        .order_by(Lotacao.area_codigo)
        .all()
    )

    proc_ids = [lot.procurador_id for lot in new_lots]
    prev_area_map = _get_prev_areas(proc_ids, ciclo_id, db)
    areas = {a.codigo: a.nome for a in db.query(Area).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Procurador", "Antiguidade", "Área Anterior", "Nome da Área Anterior",
               "Área Atual", "Nome da Área Atual", "Motivo", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
        cell.border = thin

    rows = []
    for lot in new_lots:
        proc = lot.procurador
        old_area = prev_area_map.get(lot.procurador_id)
        new_area = lot.area_codigo
        moved = old_area is None or old_area != new_area
        rows.append((
            proc.nome if proc else "",
            proc.antiguidade if proc else "",
            old_area or "—",
            areas.get(old_area, "—") if old_area else "—",
            new_area,
            areas.get(new_area, ""),
            _MOTIVO_LABEL.get(lot.motivo, lot.motivo),
            "Mudança" if moved else "Permanência",
            moved,
        ))

    rows.sort(key=lambda x: (not x[8], x[0]))

    fill_mud = PatternFill("solid", fgColor="FFFECACA")
    fill_per = PatternFill("solid", fgColor="FFBBF7D0")

    for row_idx, r in enumerate(rows, 2):
        fill = fill_mud if r[8] else fill_per
        for col, val in enumerate(r[:8], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = thin

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_decisoes_manuais_xlsx(ciclo_id: str, db: Session) -> io.BytesIO:
    import openpyxl
    PatternFill, Font, Alignment, Border, Side = _wb_styles()

    ciclo = db.get(Ciclo, ciclo_id)
    if not ciclo:
        raise ValueError(f"Ciclo '{ciclo_id}' não encontrado")

    procs = {p.id: p for p in db.query(Procurador).all()}
    areas = {a.codigo: a.nome for a in db.query(Area).all()}

    def _vagas(tipo: str, origem: Optional[str] = None):
        q = db.query(Vaga).filter(Vaga.ciclo_id == ciclo_id, Vaga.tipo == tipo)
        if origem:
            q = q.filter(Vaga.origem == origem)
        return q.order_by(Vaga.area_codigo, Vaga.numero).all()

    sheets_def = [
        ("Nomeações", _vagas("NOMEACAO"), "Nomeações da Gestão"),
        ("Escolhas de Chefes", _vagas("ESCOLHA"), "Escolhas dos Chefes"),
        ("Acervo Manual", _vagas("ACERVO", "MANUAL"), "Acervo Inserido Manualmente"),
    ]

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    headers = ["Área", "Nome da Área", "Nº Vaga", "Cargo", "Procurador", "Antiguidade"]

    wb = openpyxl.Workbook()
    first = True
    for sheet_title, vagas, tipo_label in sheets_def:
        ws = wb.active if first else wb.create_sheet(sheet_title)
        if first:
            ws.title = sheet_title
            first = False

        ws.cell(row=1, column=1, value=f"{tipo_label} — Ciclo {ciclo_id}").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
            cell.border = thin

        for row_idx, v in enumerate(vagas, 3):
            proc = procs.get(v.ocupante_id) if v.ocupante_id else None
            values = [
                v.area_codigo, areas.get(v.area_codigo, ""),
                v.numero, v.cargo or "",
                proc.nome if proc else "—",
                proc.antiguidade if proc else "",
            ]
            fill = PatternFill("solid", fgColor=_VAGA_FILL.get(v.tipo, "FFFFFFFF"))
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill = fill
                cell.border = thin

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_estatisticas_xlsx(ciclo_id: str, db: Session) -> io.BytesIO:
    import openpyxl
    PatternFill, Font, Alignment, Border, Side = _wb_styles()

    ciclo = db.get(Ciclo, ciclo_id)
    if not ciclo:
        raise ValueError(f"Ciclo '{ciclo_id}' não encontrado")

    areas_map = {a.codigo: a for a in db.query(Area).all()}

    new_lots = (
        db.query(Lotacao)
        .options(selectinload(Lotacao.procurador))
        .filter(Lotacao.ciclo_id == ciclo_id)
        .all()
    )

    proc_ids = [lot.procurador_id for lot in new_lots]
    prev_area_map = _get_prev_areas(proc_ids, ciclo_id, db)

    total = len(new_lots)
    movimentacoes = permanencias = 0
    em_regional = manteve_regional = 0
    em_esp = manteve_esp = 0

    detail_rows = []
    for lot in new_lots:
        old_cod = prev_area_map.get(lot.procurador_id)
        new_cod = lot.area_codigo
        old_area = areas_map.get(old_cod) if old_cod else None
        new_area = areas_map.get(new_cod)
        moved = old_cod is None or old_cod != new_cod

        if moved:
            movimentacoes += 1
        else:
            permanencias += 1

        if old_area and old_area.tipo == "REGIONAL":
            em_regional += 1
            if new_area and new_area.tipo == "REGIONAL":
                manteve_regional += 1

        if old_area and old_area.tipo == "ESPECIALIZADA":
            em_esp += 1
            if new_cod == old_cod:
                manteve_esp += 1

        proc = lot.procurador
        detail_rows.append((
            proc.nome if proc else "",
            proc.antiguidade if proc else "",
            old_cod or "—",
            old_area.tipo if old_area else "—",
            new_cod,
            new_area.tipo if new_area else "—",
            "Mudança" if moved else "Permanência",
        ))

    pct = lambda n, d: round(n / d * 100, 1) if d > 0 else 0.0

    metrics = [
        ("Taxa de 1ª preferência atendida (acervo)", pct(0, 1),
         ciclo.pct_primeira_pref or 0.0, "% das alocações de acervo atenderam a 1ª preferência"),
        ("Taxa de mudança geral", pct(movimentacoes, total),
         f"{movimentacoes}/{total}", "Procuradores que mudaram de área"),
        ("Taxa de permanência geral", pct(permanencias, total),
         f"{permanencias}/{total}", "Procuradores que permaneceram na mesma área"),
        ("Manutenção em regionais *", pct(manteve_regional, em_regional),
         f"{manteve_regional}/{em_regional}", "Estavam em REGIONAL e continuaram em alguma REGIONAL"),
        ("Manutenção na mesma especializada **", pct(manteve_esp, em_esp),
         f"{manteve_esp}/{em_esp}", "Estavam em ESPECIALIZADA e permaneceram na mesma área"),
    ]
    # Corrige linha 1: usa o valor do ciclo diretamente
    metrics[0] = (
        metrics[0][0],
        ciclo.pct_primeira_pref or 0.0,
        f"—/{ciclo.total_vagas or '—'}",
        metrics[0][3],
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    title_cell = ws.cell(row=1, column=1, value=f"Estatísticas — Ciclo {ciclo_id}")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    for col, h in enumerate(["Métrica", "Valor (%)", "Contagem", "Descrição"], 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
        cell.border = thin

    fills = [PatternFill("solid", fgColor="FFFFFFFF"), PatternFill("solid", fgColor="FFF9FAFB")]
    for i, (label, valor, contagem, desc) in enumerate(metrics):
        row_idx = i + 4
        f = fills[i % 2]
        for col, val in enumerate([label, f"{valor}%", contagem, desc], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = f
            cell.border = thin

    note_row = len(metrics) + 5
    ws.cell(row=note_row, column=1,
            value="* Manutenção em regionais: estava em REGIONAL antes e ficou em qualquer REGIONAL após o ciclo.")
    ws.cell(row=note_row + 1, column=1,
            value="** Manutenção na mesma especializada: permaneceu exatamente na mesma área ESPECIALIZADA.")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 55)

    # Aba de detalhes por procurador
    ws_det = wb.create_sheet("Detalhes")
    det_headers = ["Procurador", "Antiguidade", "Área Anterior", "Tipo Anterior",
                   "Área Atual", "Tipo Atual", "Status"]
    for col, h in enumerate(det_headers, 1):
        cell = ws_det.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
        cell.border = thin

    fill_mud = PatternFill("solid", fgColor="FFFECACA")
    fill_per = PatternFill("solid", fgColor="FFBBF7D0")
    for row_idx, r in enumerate(sorted(detail_rows, key=lambda x: x[0]), 2):
        fill = fill_mud if r[6] == "Mudança" else fill_per
        for col, val in enumerate(r, 1):
            cell = ws_det.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = thin

    for col in ws_det.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws_det.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_ato_pdf(ciclo_id: str, db: Session) -> io.BytesIO:
    A4, colors, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, getSS = _rl()

    ciclo = db.get(Ciclo, ciclo_id)
    if not ciclo:
        raise ValueError(f"Ciclo '{ciclo_id}' não encontrado")

    lotacoes = (
        db.query(Lotacao)
        .options(selectinload(Lotacao.procurador), selectinload(Lotacao.area))
        .filter(Lotacao.ciclo_id == ciclo_id)
        .order_by(Lotacao.area_codigo)
        .all()
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2.5*cm, rightMargin=2.5*cm,
                            topMargin=3*cm, bottomMargin=2.5*cm)
    styles = getSS()
    story = []

    hoje = date.today().strftime("%d de %B de %Y").replace(
        "January","janeiro").replace("February","fevereiro").replace("March","março")
    # Mapear nomes dos meses em inglês → português para strftime no locale padrão
    _MESES = {
        "January":"janeiro","February":"fevereiro","March":"março",
        "April":"abril","May":"maio","June":"junho","July":"julho",
        "August":"agosto","September":"setembro","October":"outubro",
        "November":"novembro","December":"dezembro",
    }
    for en, pt in _MESES.items():
        hoje = hoje.replace(en, pt)

    story.append(Paragraph("ESTADO DO RIO DE JANEIRO", styles["Normal"]))
    story.append(Paragraph("PROCURADORIA-GERAL DO ESTADO", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"<b>ATO DE RELOTAÇÃO — CICLO {ciclo_id}</b>", styles["Title"]))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"O Procurador-Geral do Estado, no uso de suas atribuições legais, "
        f"resolve relotacionar os Procuradores do Estado abaixo relacionados, "
        f"nos termos do processo de relotação periódica do ciclo <b>{ciclo_id}</b>:",
        styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    data = [["Nº", "Procurador do Estado", "Área de Lotação", "Motivo"]]
    for i, l in enumerate(lotacoes, 1):
        data.append([
            str(i),
            l.procurador.nome if l.procurador else "",
            f"{l.area_codigo} — {l.area.nome}" if l.area else l.area_codigo,
            l.motivo.replace("_", " "),
        ])

    t = Table(data, colWidths=[1*cm, 6.5*cm, 6*cm, 3*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(t)
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph(f"Rio de Janeiro, {hoje}.", styles["Normal"]))
    story.append(Spacer(1, 2*cm))
    story.append(Paragraph("_" * 40, styles["Normal"]))
    story.append(Paragraph("<b>Procurador-Geral do Estado</b>", styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return buf
