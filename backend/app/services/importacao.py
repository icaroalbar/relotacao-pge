"""
Importação de dados via planilha Excel (.xlsx).

Sheets esperadas:
  - Procuradores : Nome | Antiguidade | Status | LotacaoAtual
  - Areas        : Codigo | Nome | Tipo | VagasPG | VagasNomeacao | VagasEscolha | VagasDesignacao | VagasAcervo
  - Preferencias : coluna 0 = ProcuradorAntig, demais = área com valor = ordem (1..N)
  - Nomeacoes    : AreaCodigo | Numero | Tipo | Cargo | ProcuradorAntig  (opcional)
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional

import openpyxl
from openpyxl.workbook import Workbook


@dataclass
class ProcuradorImport:
    nome: str
    antiguidade: int
    status: str
    lotacao_atual_codigo: Optional[str]
    ativo: bool


@dataclass
class AreaImport:
    codigo: str
    nome: str
    tipo: str
    vagas_pg: int
    vagas_nomeacao: int
    vagas_escolha: int
    vagas_designacao: int
    vagas_acervo: int


@dataclass
class PreferenciaImport:
    antiguidade: int  # chave para vincular ao procurador
    area_codigo: str
    ordem: int


@dataclass
class VagaImport:
    area_codigo: str
    numero: int
    tipo: str
    cargo: Optional[str]
    procurador_antiguidade: Optional[int]


@dataclass
class ImportResult:
    procuradores: List[ProcuradorImport] = field(default_factory=list)
    areas: List[AreaImport] = field(default_factory=list)
    preferencias: List[PreferenciaImport] = field(default_factory=list)
    vagas_manuais: List[VagaImport] = field(default_factory=list)
    erros: List[str] = field(default_factory=list)


STATUS_VALIDOS = {"LOTADO", "PENDENTE", "EM_LICENCA", "VACANCIA"}
TIPO_AREA_VALIDOS = {"ESPECIALIZADA", "REGIONAL", "GABINETE"}
TIPO_VAGA_VALIDOS = {"PG", "NOMEACAO", "ESCOLHA", "DESIGNACAO", "ACERVO"}


def _cell(row, idx: int):
    val = row[idx].value
    return str(val).strip() if val is not None else None


def _int_cell(row, idx: int, default: int = 0) -> int:
    val = row[idx].value
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def ler_planilha(caminho: str) -> ImportResult:
    """
    Detecta automaticamente o formato:
    - Formato combinado: sheet com colunas Antiguidade | Nome | LotacaoAtual | [códigos de área...]
    - Formato multi-sheet: sheets separadas Procuradores + Preferencias + Areas + Nomeacoes
    """
    result = ImportResult()
    wb: Workbook = openpyxl.load_workbook(caminho, data_only=True)

    # Detectar formato combinado: primeira sheet tem "Antiguidade" e colunas de área
    primeira_sheet = wb.worksheets[0] if wb.worksheets else None
    if primeira_sheet and _is_formato_combinado(primeira_sheet):
        _ler_formato_combinado(primeira_sheet, result)
        return result

    # Formato multi-sheet tradicional
    _ler_procuradores(wb, result)
    _ler_areas(wb, result)
    _ler_preferencias(wb, result)
    _ler_nomeacoes(wb, result)

    return result


def _is_formato_combinado(ws) -> bool:
    """Detecta se a sheet tem o cabeçalho do formato combinado."""
    header = [str(c.value or "").strip().upper() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return "ANTIGUIDADE" in header and "NOME" in header


def _ler_formato_combinado(ws, result: ImportResult) -> None:
    """
    Formato único: Antiguidade | Nome | LotacaoAtual | Status(opcional) | PG-03 | PG-04 | ...
    Colunas com código de área (ex: PG-03, 1PR-NIT) são tratadas como preferências.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        result.erros.append("Planilha vazia ou sem dados")
        return

    header = [str(c or "").strip() for c in rows[0]]
    header_upper = [h.upper() for h in header]

    # Localizar colunas fixas
    try:
        col_antig = header_upper.index("ANTIGUIDADE")
    except ValueError:
        result.erros.append("Coluna 'Antiguidade' não encontrada")
        return
    try:
        col_nome = header_upper.index("NOME")
    except ValueError:
        result.erros.append("Coluna 'Nome' não encontrada")
        return

    col_lotacao = next((i for i, h in enumerate(header_upper) if "LOTA" in h), None)
    col_status  = next((i for i, h in enumerate(header_upper) if h == "STATUS"), None)

    # Colunas de preferência = cabeçalhos que parecem código de área
    import re
    _area_re = re.compile(r'^\d*[A-Z]{1,4}-?[A-Z0-9\-]+$', re.IGNORECASE)
    pref_cols: List[tuple[int, str]] = []
    for i, h in enumerate(header):
        if i in (col_antig, col_nome, col_lotacao, col_status):
            continue
        if h and _area_re.match(h.strip()):
            pref_cols.append((i, h.strip()))

    for linha_idx, row in enumerate(rows[1:], start=2):
        antig_val = row[col_antig]
        nome_val  = row[col_nome]
        if antig_val is None or nome_val is None:
            continue
        try:
            antig = int(antig_val)
        except (ValueError, TypeError):
            result.erros.append(f"Linha {linha_idx}: antiguidade inválida '{antig_val}'")
            continue

        nome = str(nome_val).strip()
        if not nome:
            continue

        lotacao = str(row[col_lotacao]).strip() if col_lotacao is not None and row[col_lotacao] else None
        status_raw = str(row[col_status]).strip().upper() if col_status is not None and row[col_status] else "PENDENTE"
        status = status_raw if status_raw in STATUS_VALIDOS else "PENDENTE"
        ativo = status not in ("EM_LICENCA", "VACANCIA")

        result.procuradores.append(ProcuradorImport(
            nome=nome, antiguidade=antig, status=status,
            lotacao_atual_codigo=lotacao or None, ativo=ativo,
        ))

        # Preferências nas colunas de área
        for col_i, area_cod in pref_cols:
            val = row[col_i] if col_i < len(row) else None
            if val is None:
                continue
            try:
                ordem = int(val)
                if ordem > 0:
                    result.preferencias.append(PreferenciaImport(
                        antiguidade=antig, area_codigo=area_cod, ordem=ordem
                    ))
            except (ValueError, TypeError):
                pass


def _ler_procuradores(wb: Workbook, result: ImportResult) -> None:
    if "Procuradores" not in wb.sheetnames:
        result.erros.append("Sheet 'Procuradores' não encontrada")
        return

    ws = wb["Procuradores"]
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        nome = _cell(row, 0)
        if not nome:
            continue
        antiguidade = _int_cell(row, 1)
        status_raw = (_cell(row, 2) or "PENDENTE").upper()
        status = status_raw if status_raw in STATUS_VALIDOS else "PENDENTE"
        if status_raw not in STATUS_VALIDOS:
            result.erros.append(f"Procuradores linha {i}: status '{status_raw}' inválido, usando PENDENTE")

        lotacao = _cell(row, 3)
        ativo = status not in ("EM_LICENCA", "VACANCIA")
        result.procuradores.append(
            ProcuradorImport(
                nome=nome,
                antiguidade=antiguidade,
                status=status,
                lotacao_atual_codigo=lotacao or None,
                ativo=ativo,
            )
        )


def _ler_areas(wb: Workbook, result: ImportResult) -> None:
    if "Areas" not in wb.sheetnames:
        result.erros.append("Sheet 'Areas' não encontrada")
        return

    ws = wb["Areas"]
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        codigo = _cell(row, 0)
        if not codigo:
            continue
        nome = _cell(row, 1) or codigo
        tipo_raw = (_cell(row, 2) or "").upper()
        tipo = tipo_raw if tipo_raw in TIPO_AREA_VALIDOS else "ESPECIALIZADA"
        if tipo_raw not in TIPO_AREA_VALIDOS:
            result.erros.append(f"Areas linha {i}: tipo '{tipo_raw}' inválido, usando ESPECIALIZADA")

        result.areas.append(
            AreaImport(
                codigo=codigo,
                nome=nome,
                tipo=tipo,
                vagas_pg=_int_cell(row, 3),
                vagas_nomeacao=_int_cell(row, 4),
                vagas_escolha=_int_cell(row, 5),
                vagas_designacao=_int_cell(row, 6),
                vagas_acervo=_int_cell(row, 7),
            )
        )


def _ler_preferencias(wb: Workbook, result: ImportResult) -> None:
    """
    Formato esperado:
      - Linha 1: cabeçalho. Célula 0 ignorada; células 1..N = códigos de área
      - Demais linhas: célula 0 = antiguidade do procurador; células 1..N = ordem (1..M) ou vazia
    """
    if "Preferencias" not in wb.sheetnames:
        return

    ws = wb["Preferencias"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return

    header = rows[0]
    codigos = [str(c).strip() if c else None for c in header[1:]]

    for row in rows[1:]:
        try:
            antig = int(row[0])
        except (TypeError, ValueError):
            continue
        for col_idx, ordem_val in enumerate(row[1:]):
            if ordem_val is None:
                continue
            try:
                ordem = int(ordem_val)
            except (TypeError, ValueError):
                continue
            area_codigo = codigos[col_idx]
            if area_codigo:
                result.preferencias.append(
                    PreferenciaImport(antiguidade=antig, area_codigo=area_codigo, ordem=ordem)
                )


def _ler_nomeacoes(wb: Workbook, result: ImportResult) -> None:
    """
    Sheet opcional. Colunas: AreaCodigo | Numero | Tipo | Cargo | ProcuradorAntig
    """
    for sheet_name in ("Nomeacoes", "Nomeações", "Designacoes", "Escolhas"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            area_codigo = _cell(row, 0)
            if not area_codigo:
                continue
            numero = _int_cell(row, 1, default=1)
            tipo_raw = (_cell(row, 2) or "NOMEACAO").upper()
            tipo = tipo_raw if tipo_raw in TIPO_VAGA_VALIDOS else "NOMEACAO"
            cargo = _cell(row, 3)
            antig_raw = row[4].value if len(row) > 4 else None
            try:
                proc_antig = int(antig_raw) if antig_raw is not None else None
            except (TypeError, ValueError):
                proc_antig = None

            result.vagas_manuais.append(
                VagaImport(
                    area_codigo=area_codigo,
                    numero=numero,
                    tipo=tipo,
                    cargo=cargo,
                    procurador_antiguidade=proc_antig,
                )
            )
